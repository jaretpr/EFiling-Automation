import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import scrolledtext
import time
from datetime import datetime

# Mapping of attorney initials to login credentials
credentials = {
    "ATTORNEY1": {"user": "USERNAME1", "pass": "PASSWORD1"},
    "ATTORNEY2": {"user": "USERNAME2", "pass": "PASSWORD2"},
    "ATTORNEY3": {"user": "USERNAME3", "pass": "PASSWORD3"},
    # Add more as needed
}

def get_court_value(court_name):
    court_dict = {
        "Allen County District Court": "123110",
        "Anderson County District Court": "120410",
        "Atchison County District Court": "120110",
        "Barber County District Court": "123010",
        "Barton County District Court": "122010",
        "Bourbon County District Court": "120610",
        "Brown County District Court": "122210",
        "Butler County District Court": "121310",
        "Chase County District Court": "120510",
        "Chautauqua County District Court": "121410",
        "Cherokee County District Court": "121110",
        "Cheyenne County District Court": "121510",
        "Clark County District Court": "121610",
        "Clay County District Court": "122110",
        "Cloud County District Court": "121210",
        "Coffey County District Court": "120420",
        "Comanche County District Court": "121620",
        "Cowley-Arkansas City District Court": "1219110",
        "Cowley-Winfield District Court": "1219120",
        "Crawford-Girard District Court": "1211210",
        "Crawford-Pittsburg District Court": "1211220",
        "Decatur County District Court": "121710",
        "Dickinson County District Court": "120810",
        "Doniphan County District Court": "122220",
        "Douglas County District Court": "120710",
        "Edwards County District Court": "122410",
        "Elk County District Court": "121320",
        "Ellis County District Court": "122310",
        "Ellsworth County District Court": "122020",
        "Finney County District Court": "122510",
        "Ford County District Court": "121630",
        "Franklin County District Court": "120430",
        "Geary County District Court": "120820",
        "Gove County District Court": "122320",
        "Graham County District Court": "121720",
        "Grant County District Court": "122610",
        "Gray County District Court": "121640",
        "Greeley County District Court": "122520",
        "Greenwood County District Court": "121330",
        "Hamilton County District Court": "122530",
        "Harper County District Court": "123020",
        "Harvey County District Court": "120910",
        "Haskell County District Court": "122620",
        "Hodgeman County District Court": "122420",
        "Jackson County District Court": "120210",
        "Jefferson County District Court": "120220",
        "Jewell County District Court": "121220",
        "Kearny County District Court": "122540",
        "Kingman County District Court": "123030",
        "Kiowa County District Court": "121650",
        "Labette-Oswego District Court": "1211310",
        "Labette-Parsons District Court": "1211320",
        "Lane County District Court": "122430",
        "Leavenworth County District Court": "120120",
        "Lincoln County District Court": "121230",
        "Linn County District Court": "120620",
        "Logan County District Court": "121520",
        "Lyon County District Court": "120520",
        "Marion County District Court": "120830",
        "Marshall County District Court": "122230",
        "McPherson County District Court": "120920",
        "Meade County District Court": "121660",
        "Miami County District Court": "120630",
        "Mitchell County District Court": "121240",
        "Montgomery-Coffeyville District Court": "1214210",
        "Montgomery-Independence District Court": "1214220",
        "Morris County District Court": "120840",
        "Morton County District Court": "122630",
        "Nemaha County District Court": "122240",
        "Neosho-Chanute District Court": "1231210",
        "Neosho-Erie District Court": "1231220",
        "Ness County District Court": "122440",
        "Norton County District Court": "121730",
        "Osage County District Court": "120440",
        "Osborne County District Court": "121740",
        "Ottawa County District Court": "122810",
        "Pawnee County District Court": "122450",
        "Phillips County District Court": "121750",
        "Pottawatomie County District Court": "120230",
        "Pratt County District Court": "123040",
        "Rawlins County District Court": "121530",
        "Reno County District Court": "122710",
        "Republic County District Court": "121250",
        "Rice County District Court": "122030",
        "Riley County District Court": "122120",
        "Rooks County District Court": "122330",
        "Rush County District Court": "122460",
        "Russell County District Court": "122040",
        "Saline County District Court": "122820",
        "Scott County District Court": "122550",
        "Sedgwick County District Court": "121810",
        "Seward County District Court": "122640",
        "Shawnee County District Court": "120310",
        "Sheridan County District Court": "121540",
        "Sherman County District Court": "121550",
        "Smith County District Court": "121760",
        "Stafford County District Court": "122050",
        "Stanton County District Court": "122650",
        "Stevens County District Court": "122660",
        "Sumner County District Court": "123050",
        "Thomas County District Court": "121560",
        "Trego County District Court": "122340",
        "Wabaunsee County District Court": "120240",
        "Wallace County District Court": "121570",
        "Washington County District Court": "121260",
        "Wichita County District Court": "122560",
        "Wilson County District Court": "123130",
        "Woodson County District Court": "123140",
        "Wyandotte County District Court": "122910",
    }

    # Special cases for specific prefixes
    prefix_court_mapping = {
        "CLA - Cowley County District Court": "Cowley-Arkansas City District Court",
        "CLW - Cowley County District Court": "Cowley-Winfield District Court",
        "CRP - Crawford County District Court": "Crawford-Pittsburg District Court",
        "LBP - Labette County District Court": "Labette-Parsons District Court",
        "LBO - Labette County District Court": "Labette-Oswego District Court",
        "MGI - Montgomery County District Court": "Montgomery-Independence District Court",
        "MGC - Montgomery County District Court": "Montgomery-Coffeyville District Court",
        "CRG - Crawford County District Court": "Crawford-Girard District Court",
        "NOC - Neosho County District Court": "Neosho-Chanute District Court",
        "NOE - Neosho County District Court": "Neosho-Erie District Court"
    }

    for prefix in prefix_court_mapping:
        if court_name.startswith(prefix):
            specific_court_name = prefix_court_mapping[prefix]
            return court_dict.get(specific_court_name, "Unknown Court")

    if "Shawnee County Limited Actions District Court" in court_name:
        return court_dict["Shawnee County District Court"]

    prefixes = [
    'AL - ', 'GW - ', 'OS - ', 'AN - ', 'GY - ', 'OT - ', 'AT - ', 'HG - ', 'PL - ', 
    'BA - ', 'HM - ', 'PN - ', 'BB - ', 'HP - ', 'PR - ', 'BR - ', 'HS - ', 'PT - ', 
    'BT - ', 'HV - ', 'RA - ', 'BU - ', 'JA - ', 'RC - ', 'CA - ', 'JF - ', 'RH - ', 
    'CD - ', 'JO - ', 'RL - ', 'CF - ', 'JW - ', 'RN - ', 'CK - ', 'KE - ', 'RO - ', 
    'CL - ', 'KM - ', 'RP - ', 'CM - ', 'KW - ', 'RS - ', 'CN - ', 'LB - ', 'SA - ', 
    'CQ - ', 'LC - ', 'SC - ', 'CR - ', 'LE - ', 'SD - ', 'CS - ', 'LG - ', 'SF - ', 
    'CY - ', 'LN - ', 'SG - ', 'DC - ', 'LV - ', 'SH - ', 'DG - ', 'LY - ', 'SM - ', 
    'DK - ', 'MC - ', 'SN - ', 'DP - ', 'ME - ', 'ST - ', 'ED - ', 'MG - ', 'SU - ', 
    'EK - ', 'MI - ', 'SV - ', 'EL - ', 'MN - ', 'SW - ', 'EW - ', 'MP - ', 'TH - ', 
    'FI - ', 'MR - ', 'TR - ', 'FO - ', 'MS - ', 'WA - ', 'FR - ', 'MT - ', 'WB - ', 
    'GE - ', 'NM - ', 'WH - ', 'GH - ', 'NO - ', 'WL - ', 'GL - ', 'NS - ', 'WO - ', 
    'GO - ', 'NT - ', 'WS - ', 'GT - ', 'OB - ', 'WY - ', 'SNCO - ', 'DP - ', 'BR - ',
    'FR - ', 'JA - ', 'LY - ', 'RC - ', 'RN - ', 'SG - ', 'WY - ', 'DG - ', 'RL - ',
    'SF - ', 'JF - ', 'CQ - ', 'PT - ', 'RS - ', 'SA - ', 'CS - ', 'FO - ', 'WB - ', 
    'PR - ', 'HV - ', 'FI - ', 'DK - ', 'WL - ', 'BA - ', 'CK - ', 'CS - ', 'CY - ', 
    'EL - ', 'KM - ', 'LN - ', 'MP - ', 'PN - ', 'SW - ', 'WO - ', 'EK - ', 'LV - ', 
    'MR - ', 'MS - ', 'NM - '
]

    cleaned_court_name = court_name.strip()
    for prefix in prefixes:
        if cleaned_court_name.startswith(prefix):
            cleaned_court_name = cleaned_court_name.replace(prefix, '').strip()

    return court_dict.get(cleaned_court_name, "Unknown Court")

def strip_prefix(filers_case_number):
    if filers_case_number is None:
        return filers_case_number
        prefixes = [
    'AL - ', 'GW - ', 'OS - ', 'AN - ', 'GY - ', 'OT - ', 'AT - ', 'HG - ', 'PL - ', 
    'BA - ', 'HM - ', 'PN - ', 'BB - ', 'HP - ', 'PR - ', 'BR - ', 'HS - ', 'PT - ', 
    'BT - ', 'HV - ', 'RA - ', 'BU - ', 'JA - ', 'RC - ', 'CA - ', 'JF - ', 'RH - ', 
    'CD - ', 'JO - ', 'RL - ', 'CF - ', 'JW - ', 'RN - ', 'CK - ', 'KE - ', 'RO - ', 
    'CL - ', 'KM - ', 'RP - ', 'CM - ', 'KW - ', 'RS - ', 'CN - ', 'LB - ', 'SA - ', 
    'CQ - ', 'LC - ', 'SC - ', 'CR - ', 'LE - ', 'SD - ', 'CS - ', 'LG - ', 'SF - ', 
    'CY - ', 'LN - ', 'SG - ', 'DC - ', 'LV - ', 'SH - ', 'DG - ', 'LY - ', 'SM - ', 
    'DK - ', 'MC - ', 'SN - ', 'DP - ', 'ME - ', 'ST - ', 'ED - ', 'MG - ', 'SU - ', 
    'EK - ', 'MI - ', 'SV - ', 'EL - ', 'MN - ', 'SW - ', 'EW - ', 'MP - ', 'TH - ', 
    'FI - ', 'MR - ', 'TR - ', 'FO - ', 'MS - ', 'WA - ', 'FR - ', 'MT - ', 'WB - ', 
    'GE - ', 'NM - ', 'WH - ', 'GH - ', 'NO - ', 'WL - ', 'GL - ', 'NS - ', 'WO - ', 
    'GO - ', 'NT - ', 'WS - ', 'GT - ', 'OB - ', 'WY - ', 'SNCO - ', 'DP - ', 'BR - ',
    'FR - ', 'JA - ', 'LY - ', 'RC - ', 'RN - ', 'SG - ', 'WY - ', 'DG - ', 'RL - ',
    'SF - ', 'JF - ', 'CQ - ', 'PT - ', 'RS - ', 'SA - ', 'CS - ', 'FO - ', 'WB - ', 
    'PR - ', 'HV - ', 'FI - ', 'DK - ', 'WL - ', 'BA - ', 'CK - ', 'CS - ', 'CY - ', 
    'EL - ', 'KM - ', 'LN - ', 'MP - ', 'PN - ', 'SW - ', 'WO - ', 'EK - ', 'LV - ', 
    'MR - ', 'MS - ', 'NM - '
]

    for prefix in prefixes:
        if filers_case_number.startswith(prefix):
            return filers_case_number.replace(prefix, '').strip()
    return filers_case_number
    
def format_document_title(note, court_name):
    if "Shawnee County District Court" in court_name:
        return "Return of Service"
    cleaned_note = re.sub(r"\[.*?\]", "", note).strip()
    if 'Wage' in cleaned_note:
        return "Return of Service of Wage Garnishment"
    elif 'Bank' in cleaned_note:
        return "Return of Service of Non-Wage Garnishment"
    else:
        return "Return of Service"

def clean_case_number(case_number):
    if case_number is None:
        return case_number
    match = re.match(r"(\d{4}-[A-Z]{2}-\d{6})(-[A-Z]{2})?", case_number)
    if match:
        return match.group(1)
    return case_number

def start_efiling():
    log_widget.insert(tk.END, "Starting e-filing process...\n")
    
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    service = Service(executable_path="path/to/chromedriver")
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    driver.get("https://example.com/portal/")
    wait = WebDriverWait(driver, 20)
    
    current_date = datetime.now().strftime('%m.%d.%y')
    wb_path = fr'path/to/excel/file/{current_date} Garn Returns.xlsx'
    
    if not os.path.exists(wb_path):
        log_widget.insert(tk.END, f"File not found: {wb_path}\n")
        driver.quit()
        return

    try:
        wb = load_workbook(wb_path)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            court_file_no = row[0].value
            debtor_number = row[1].value
            court_name = row[3].value
            note = row[4].value
            completed = row[5]
            attorney_initials = row[6].value

            if completed.value == "Yes":
                continue

            if court_file_no is None or debtor_number is None or court_name is None or note is None:
                log_widget.insert(tk.END, f"Skipping row with missing data: {row}\n")
                continue
                
            if attorney_initials in credentials:
                user = credentials[attorney_initials]["user"]
                password = credentials[attorney_initials]["pass"]
            else:
                log_widget.insert(tk.END, f"Unknown attorney initials: {attorney_initials}\n")
                continue
                
            wait.until(EC.presence_of_element_located((By.ID, "userName"))).send_keys(user)
            wait.until(EC.presence_of_element_located((By.ID, "password"))).send_keys(password)
            wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input.sbttn[value='Log In']"))).click()

            cleaned_case_number = clean_case_number(strip_prefix(court_file_no))

            wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Cases"))).click()
            wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "My Cases"))).click()

            wait.until(EC.presence_of_element_located((By.ID, "caseNumber"))).send_keys(cleaned_case_number)

            court_value = get_court_value(court_name)
            if court_value == "Unknown Court":
                log_widget.insert(tk.END, f"Unknown court: {court_name}\n")
                continue

            court_select = Select(wait.until(EC.presence_of_element_located((By.ID, "wkLocation"))))
            court_select.select_by_value(court_value)

            wait.until(EC.element_to_be_clickable((By.NAME, "eFile"))).click()

            category_select = Select(wait.until(EC.element_to_be_clickable((By.ID, "categories"))))
            category_select.select_by_value("121")

            doc_type_select = Select(wait.until(EC.element_to_be_clickable((By.ID, "ftDocDefId"))))
            found = False
            for option in doc_type_select.options:
                if "Return of Service" in option.text:
                    doc_type_select.select_by_visible_text(option.text)
                    found = True
                    break
            if not found:
                log_widget.insert(tk.END, "Document type for Return of Service not found.\n")
                driver.quit()
                return

            title = format_document_title(note, court_name)
            wait.until(EC.presence_of_element_located((By.ID, "postDescription"))).send_keys(title)

            file_type_prefix = 'GJ' if 'bank' in note.lower() else 'GI'
            file_name = f"{file_type_prefix}{int(debtor_number):010}.pdf"
            file_path = os.path.join(r"path/to/pdf/files", file_name)
            
            if not os.path.exists(file_path):
                file_name = f"{file_type_prefix}{int(debtor_number):010}_001.pdf"
                file_path = os.path.join(r"path/to/pdf/files", file_name)
                
                if not os.path.exists(file_path):
                    log_widget.insert(tk.END, f"File not found: {file_path}\n")
                    continue
                    
            wait.until(EC.presence_of_element_located((By.ID, "Filename"))).send_keys(os.path.abspath(file_path))
            wait.until(EC.element_to_be_clickable((By.ID, "addButton"))).click()
            
            wait.until(EC.element_to_be_clickable((By.ID, "nextButton"))).click()
            
            stripped_debtor_number = debtor_number.strip()

            wait.until(EC.presence_of_element_located((By.ID, "FilersCaseNumber"))).send_keys(stripped_debtor_number)
            
            wait.until(EC.presence_of_element_located((By.NAME, "piiCheckBox"))).click()
            
            wait.until(EC.presence_of_element_located((By.ID, "submitButton"))).click()
            
            WebDriverWait(driver, 10).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            alert.accept()

            completed.value = "Yes"
            wb.save(wb_path)
            
            log_widget.insert(tk.END, f"Processed document for case: {cleaned_case_number}\n")

            driver.get("https://example.com/portal/")
            wait.until(EC.presence_of_element_located((By.ID, "userName")))

    except Exception as e:
        log_widget.insert(tk.END, f"An error occurred: {str(e)}\n")
    
    finally:
        driver.quit()
        log_widget.insert(tk.END, "E-filing process completed.\n")

# Tkinter GUI setup
root = tk.Tk()
root.title("Garnishment Service Returns")
root.geometry("600x600")
root.configure(bg='#2d2d30')

frame = tk.Frame(root, padx=20, pady=20, bg='#2d2d30')
frame.pack(expand=True, fill=tk.BOTH)

title_label = tk.Label(frame, text="File Uploader", font=("Arial", 16, "bold"), fg="white", bg="#2d2d30")
title_label.grid(row=0, column=0, columnspan=2, pady=10, sticky="n")

button_style = {"font": ("Arial", 12), "padx": 5, "pady": 5, "bg": "#1c97ea", "fg": "white", "relief": tk.RAISED, "borderwidth": 2}
log_style = {"bg": "#1e1e1e", "fg": "white", "borderwidth": 2, "relief": tk.RAISED}

btn_upload = tk.Button(frame, text="Start Process", command=start_efiling, **button_style)
btn_upload.grid(row=1, column=0, pady=5, padx=10, sticky="ew")

log_widget = scrolledtext.ScrolledText(frame, wrap=tk.WORD, height=20, font=("Arial", 10), **log_style)
log_widget.grid(row=2, column=0, columnspan=2, pady=5, padx=5, sticky="nsew")

frame.grid_columnconfigure(0, weight=1)
frame.grid_rowconfigure(2, weight=1)

root.mainloop()
